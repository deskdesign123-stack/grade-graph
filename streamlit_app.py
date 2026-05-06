import streamlit as st
import openpyxl
from PIL import Image, ImageDraw, ImageFont
import re, io, zipfile

# ── 폰트
F_EB = "/usr/share/fonts/truetype/nanum/NanumSquareEB.ttf"
F_B  = "/usr/share/fonts/truetype/nanum/NanumSquareB.ttf"

W, H  = 1754, 1241
SCALE = 1754 / 865

BLUE     = (21,  115, 185)
GRAY_BAR = (172, 168, 167)
GRAY_TXT = (90,   86,  85)
RED_TXT  = (204,  17,  60)
WHITE    = (255, 255, 255)
BLACK    = (30,   30,  30)
LINE_CLR = (120, 120, 120)


def parse_name(raw):
    s = re.sub(r'^\([^)]*\)', '', str(raw)).strip()
    s = re.sub(r'\d+$', '', s).strip()
    return s[:3] if len(s) >= 3 else s

def mask_name(name):
    if len(name) == 2: return name[0] + 'O'
    if len(name) >= 3: return name[0] + 'O' + name[2:]
    return name

def lf(path, size): return ImageFont.truetype(path, size)
def tw(draw, t, f): return draw.textlength(t, font=f)


def compute_heights(subjects, blue_max):
    # 8등급은 상대 스케일 계산에서 제외 (다른 과목 막대 왜곡 방지)
    prev_gs = [s['prev_grade'] for s in subjects
               if s['prev_grade'] is not None and s['prev_grade'] < 8]
    curr_gs = [s['curr_grade'] for s in subjects if s['curr_grade'] is not None]

    def to_h(g, gs, lo, hi, area):
        if g is None or not gs: return int((lo+hi)/2 * area)
        mn, mx = min(gs), max(gs)
        if mn == mx: return int((lo+hi)/2 * area)
        ratio = (mx - g) / (mx - mn)
        return int((lo + ratio*(hi-lo)) * area)

    result = []
    for s in subjects:
        pg, cg = s['prev_grade'], s['curr_grade']
        diff = (pg - cg) if (pg is not None and cg is not None) else 3

        # 8등급은 항상 고정 최저 높이
        if pg == 8:
            ph = int(blue_max * 0.12)
        else:
            if cg is not None and cg >= 3:
                if diff <= 3:   gray_lo, gray_hi = 0.36, 0.52
                else:           gray_lo, gray_hi = 0.20, 0.38
            else:
                if diff <= 2:   gray_lo, gray_hi = 0.52, 0.67
                elif diff == 3: gray_lo, gray_hi = 0.36, 0.52
                else:           gray_lo, gray_hi = 0.20, 0.38
            ph = to_h(pg, prev_gs, gray_lo, gray_hi, blue_max)

        ch = to_h(cg, curr_gs, 0.72, 0.93, blue_max)
        min_gap = int(blue_max * 0.08)
        if ch <= ph + min_gap: ch = ph + min_gap
        ch = min(ch, int(blue_max * 0.93))
        result.append((ph, ch))
    return result


def draw_student(name_raw, subjects):
    """이미지 생성 후 bytes 반환"""
    name3  = parse_name(name_raw)
    masked = mask_name(name3)

    img  = Image.new('RGB', (W, H), WHITE)
    draw = ImageDraw.Draw(img)

    fn_title = lf(F_EB, 80)
    fn_red   = lf(F_EB, 46)
    fn_gray  = lf(F_EB, 40)
    fn_subj  = lf(F_EB, 48)
    fn_note  = lf(F_B,  28)

    name_disp = masked + " 학생"
    nw = tw(draw, name_disp, fn_title)
    draw.text(((W - nw)//2, int(22*SCALE)), name_disp, fill=BLACK, font=fn_title)

    CL       = int(27  * SCALE)
    CR       = int(843 * SCALE)
    BASE_Y   = int(493 * SCALE)
    BLUE_MAX = int(336 * SCALE)

    n = len(subjects)
    if n == 0:
        buf = io.BytesIO()
        img.save(buf, format='PNG', dpi=(150,150))
        return buf.getvalue()

    col_w   = (CR - CL) // n
    BW_BLUE = int(col_w * 0.385)
    BW_GRAY = int(col_w * 0.270)
    GAP     = int(col_w * 0.046)

    heights = compute_heights(subjects, BLUE_MAX)

    red_bb  = fn_red.getbbox("1등급")
    red_h   = red_bb[3] - red_bb[1]
    RED_GAP = int(11 * SCALE)

    for i, (subj, (ph, ch)) in enumerate(zip(subjects, heights)):
        cx     = CL + i * col_w + col_w // 2
        prev_g = subj['prev_grade']
        curr_g = subj['curr_grade']
        same   = (prev_g is not None and curr_g is not None and prev_g == curr_g)
        sname  = subj['name']

        if same:
            bx = cx - BW_BLUE // 2
            by = BASE_Y - ch
            draw.rectangle([bx, by, bx+BW_BLUE, BASE_Y], fill=BLUE)
            gl = f"{curr_g}등급"
            draw.text((bx+(BW_BLUE-tw(draw,gl,fn_red))//2, by-RED_GAP-red_h), gl, fill=RED_TXT, font=fn_red)
        else:
            if prev_g is not None:
                bx = cx - BW_GRAY - GAP//2
                by = BASE_Y - ph
                draw.rectangle([bx, by, bx+BW_GRAY, BASE_Y], fill=GRAY_BAR)
                gl  = f"{prev_g}등급"
                glw = tw(draw, gl, fn_gray)
                gray_th = fn_gray.getbbox(gl)[3] - fn_gray.getbbox(gl)[1]
                draw.text((bx+(BW_GRAY-glw)//2, by-int(6*SCALE)-gray_th), gl, fill=GRAY_TXT, font=fn_gray)
            if curr_g is not None:
                bx = cx + GAP//2
                by = BASE_Y - ch
                draw.rectangle([bx, by, bx+BW_BLUE, BASE_Y], fill=BLUE)
                gl = f"{curr_g}등급"
                draw.text((bx+(BW_BLUE-tw(draw,gl,fn_red))//2, by-RED_GAP-red_h), gl, fill=RED_TXT, font=fn_red)

        snw = tw(draw, sname, fn_subj)
        draw.text((cx-snw//2, BASE_Y+int(14*SCALE)), sname, fill=BLACK, font=fn_subj)

    draw.line([(CL, BASE_Y), (CR, BASE_Y)], fill=LINE_CLR, width=3)

    if any(s.get('needs_note') for s in subjects):
        note_txt = "*수능 미응시 등급 8등급으로 환산"
        ntw = tw(draw, note_txt, fn_note)
        draw.text((CR-ntw, BASE_Y+int(75*SCALE)), note_txt, fill=GRAY_TXT, font=fn_note)

    buf = io.BytesIO()
    img.save(buf, format='PNG', dpi=(150,150))
    return buf.getvalue()


def is_red(cell):
    try: return cell.fill.fgColor.rgb == 'FFFF0000'
    except: return False

def get_int(cell):
    v = cell.value
    if v is None or str(v).strip() == '': return None
    try: return int(v)
    except: return None

def get_str(cell):
    v = cell.value
    if v is None or str(v).strip() == '': return None
    return str(v).strip()

def load_students(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    students = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        if not row[2].value or str(row[2].value).strip() == '': continue
        name_raw = str(row[2].value).strip()

        def fixed(label, pi, ci):
            pg, cg = row[pi], row[ci]
            if is_red(pg) and is_red(cg): return None
            pv = None if is_red(pg) else get_int(pg)
            cv = None if is_red(cg) else get_int(cg)
            if pv is None and cv is None: return None
            return {'name': label, 'prev_grade': pv, 'curr_grade': cv, 'needs_note': False}

        def flexible(ps, pg, cs, cg):
            p_subj_red  = is_red(row[ps])
            p_grade_red = is_red(row[pg])
            c_subj_red  = is_red(row[cs])
            c_grade_red = is_red(row[cg])
            if (p_subj_red or p_grade_red) and (c_subj_red or c_grade_red): return None
            curr_g    = None if (c_subj_red or c_grade_red) else get_int(row[cg])
            curr_name = get_str(row[cs])
            prev_name = get_str(row[ps])
            prev_g    = None if (p_subj_red or p_grade_red) else get_int(row[pg])
            if curr_g is None and prev_g is None: return None
            name = curr_name or prev_name or '탐구'
            needs_note = False
            if p_subj_red or p_grade_red:
                prev_g = 8; needs_note = True
            elif prev_name and curr_name and prev_name.replace(' ','') != curr_name.replace(' ',''):
                prev_g = 8; needs_note = True
            return {'name': name, 'prev_grade': prev_g, 'curr_grade': curr_g, 'needs_note': needs_note}

        subs = []
        e  = fixed('국어', 5, 14);    
        if e:  subs.append(e)
        s  = fixed('수학', 7, 16);    
        if s:  subs.append(s)
        g  = fixed('영어', 8, 17);    
        if g:  subs.append(g)
        t1 = flexible(9, 10, 18, 19)
        if t1: subs.append(t1)
        t2 = flexible(11, 12, 20, 21)
        if t2: subs.append(t2)
        if subs: students.append((name_raw, subs))
    return students


# ══════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════
st.set_page_config(page_title="성적 향상 그래프 생성기", page_icon="📊", layout="centered")

st.title("📊 성적 향상 그래프 생성기")
st.caption("엑셀 파일을 업로드하면 학생별 성적 향상 이미지를 자동으로 만들어드립니다.")
st.divider()

uploaded = st.file_uploader("📂 엑셀 파일 업로드 (.xlsx)", type=["xlsx"])

if uploaded:
    st.success(f"✅ 파일 업로드 완료: **{uploaded.name}**")

    if st.button("🖼  이미지 생성 시작", type="primary", use_container_width=True):

        with st.spinner("엑셀 파일 읽는 중..."):
            students = load_students(uploaded.read())

        st.info(f"총 **{len(students)}명** 발견. 이미지를 생성합니다...")

        progress = st.progress(0, text="준비 중...")
        zip_buf  = io.BytesIO()

        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for i, (name_raw, subjects) in enumerate(students):
                name3 = parse_name(name_raw)
                safe  = re.sub(r'[^\w가-힣]', '_', name3)
                fname = f"{i+1:02d}_{safe}.png"

                progress.progress((i+1)/len(students),
                                  text=f"생성 중... ({i+1}/{len(students)}) {name3}")

                img_bytes = draw_student(name_raw, subjects)
                zf.writestr(fname, img_bytes)

        progress.progress(1.0, text="✅ 완료!")
        st.balloons()
        st.success(f"🎉 **{len(students)}명** 이미지 생성 완료!")

        st.download_button(
            label="⬇️ ZIP 다운로드",
            data=zip_buf.getvalue(),
            file_name="성적향상_그래프.zip",
            mime="application/zip",
            use_container_width=True,
            type="primary"
        )

        # 미리보기 (첫 번째 학생)
        with st.expander("🔍 첫 번째 학생 미리보기"):
            first_name, first_subs = students[0]
            st.image(draw_student(first_name, first_subs),
                     caption=f"{parse_name(first_name)} 학생", use_container_width=True)
